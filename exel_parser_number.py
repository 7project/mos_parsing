from openpyxl import load_workbook

data_numbers = []


def xlsx_data_parsing():
    wb = load_workbook('./data.xlsx')
    sheet = wb.worksheets[1]

    for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row):
        if row[1].value:
            data_numbers.append(row[1].value)

    print('Load finish list')
    print(len(data_numbers))

    wb.close()
    return data_numbers


def one_number_get(numbers):
    for number in numbers:
        yield number

